import requests
from datetime import datetime
import os
from collections import Counter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# ====== GitHub API-uppsättning ======
GITHUB_API_URL = "https://api.github.com/search/users"
HEADERS = {
    "Accept": "application/vnd.github+json",
    "Authorization": f"Bearer {os.getenv('GITHUB_TOKEN')}"
}

# === Keywordlistor ===
CONSULTANT_KEYWORDS = [
    "consultant", "consulting", "konsult", "konsulter", "contractor", "external consultant",
    "technical consultant", "utvecklingskonsult", "systemkonsult", "software consultant"
]
FREELANCE_KEYWORDS = [
    "freelancer", "freelance", "freelancing", "self-employed", "egenföretagare",
    "own business", "entrepreneur", "frilans", "frilansare", "independent", "solopreneur",
    "gig worker", "contract work", "open for freelance"
]
CONSULTANCY_COMPANIES = [
    "Netlight", "HiQ", "Knowit", "Cybercom", "Cygni", "Acando", "TIQQE", "Softronic",
    "Sigma", "Columbus", "Sogeti", "Avega", "Deloitte", "Zington", "Tretton37",
    "Capgemini", "CGI", "Xlent", "AFRY", "ALTEN", "EVRY", "Consid", "B3",
    "QGroup", "Devoteam", "Northab", "Centigo", "Combitech", "Ninetech", "AddQ"
]
COMPANY_HINTS = [".com", "AB", "Tech", "dev", "company", "consulting"]

LANGUAGE_ORDER = ["Java", "Kotlin", ".Net", "Go", "Python", "JavaScript", "Annat"]

def fetch_profiles(languages, location="Stockholm", max_results=60, mode="all"):
    results = []
    per_page = 30
    max_pages = (max_results // per_page) if max_results != -1 else 10

    for language in languages:
        for page in range(1, max_pages + 1):
            query = f"location:{location} language:{language} repos:>0"
            params = {
                "q": query,
                "per_page": per_page,
                "page": page
            }
            response = requests.get(GITHUB_API_URL, headers=HEADERS, params=params)
            if response.status_code != 200:
                break
            items = response.json().get("items", [])

            for user in items:
                profile = parse_profile(user["url"], languages, mode, location)
                if profile:
                    results.append(profile)
                    if max_results != -1 and len(results) >= max_results:
                        return results

    return results

def parse_profile(api_url, languages, mode, required_location):
    res = requests.get(api_url, headers=HEADERS)
    if res.status_code != 200:
        return None

    data = res.json()
    bio = (data.get("bio") or "")
    company = (data.get("company") or "")
    location = (data.get("location") or "")
    repos = data.get("public_repos", 0)
    created_at = data.get("created_at", "")
    created_year = int(created_at[:4]) if created_at else 0
    html_url = data.get("html_url", "")
    username = data.get("login", "")

    langs_count = count_languages(username)
    top_langs = ", ".join([f"{lang}: {count}" for lang, count in langs_count.items() if count > 0])
    primary_lang = max(langs_count, key=langs_count.get) if langs_count else "Annat"

    metadata_text = " ".join([bio.lower(), company.lower()])
    is_consultant = any(word in metadata_text for word in CONSULTANT_KEYWORDS)
    is_freelancer = any(word in metadata_text for word in FREELANCE_KEYWORDS)
    is_consulting_company = any(bolag.lower() in company.lower() for bolag in CONSULTANCY_COMPANIES)
    company_hint = any(hint.lower() in metadata_text for hint in COMPANY_HINTS)
    created_old_enough = created_year <= datetime.now().year - 6

    # === Whitelist-läge ===
    if required_location.lower() not in location.lower():
        return None

    if mode == "only_consultants":
        if is_consultant or is_freelancer or is_consulting_company:
            return _build_profile(html_url, bio, repos, created_year, top_langs, primary_lang)
        return None

    if mode == "for_employment":
        if is_freelancer and not company_hint:
            return None
        if not created_old_enough and not company_hint:
            return None
        return _build_profile(html_url, bio, repos, created_year, top_langs, primary_lang)

    # mode == all
    if not created_old_enough and not company_hint:
        return None
    return _build_profile(html_url, bio, repos, created_year, top_langs, primary_lang)

def _build_profile(html_url, bio, repos, created_year, top_langs, primary_lang):
    return {
        "GitHub-profil": html_url,
        "Bio": bio,
        "Publika repos": repos,
        "År skapad": created_year,
        "Språk (antal repos)": top_langs,
        "Primärt språk": primary_lang
    }

def count_languages(username):
    url = f"https://api.github.com/users/{username}/repos"
    langs = {}
    page = 1
    while True:
        params = {"per_page": 100, "page": page}
        res = requests.get(url, headers=HEADERS, params=params)
        if res.status_code != 200:
            break
        repos = res.json()
        if not repos:
            break
        for repo in repos:
            lang = repo.get("language")
            if lang:
                langs[lang] = langs.get(lang, 0) + 1
        page += 1
    return langs

def sort_profiles(profiles):
    def sort_key(profile):
        lang = profile.get("Primärt språk", "Annat")
        return LANGUAGE_ORDER.index(lang) if lang in LANGUAGE_ORDER else len(LANGUAGE_ORDER)
    return sorted(profiles, key=sort_key)

def save_to_excel(profiles, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "GitHub-profiler"

    headers = ["GitHub-profil", "Bio", "Publika repos", "År skapad", "Språk (antal repos)", "Primärt språk"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for profile in profiles:
        row = [
            profile["GitHub-profil"],
            profile["Bio"],
            profile["Publika repos"],
            profile["År skapad"],
            profile["Språk (antal repos)"],
            profile["Primärt språk"]
        ]
        ws.append(row)

    for row in range(2, len(profiles) + 2):
        url = ws[f"A{row}"].value
        ws[f"A{row}"].hyperlink = url
        ws[f"A{row}"].style = "Hyperlink"

    wb.save(path)
